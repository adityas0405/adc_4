#Phyton file will be described to calibrate real scenerio of temperature data to an ideal set of temperature data 
#all data extracted from a temerature sensor using a bjt), method of calibration used will be one-point calibration
#data will be stored and displayed on an excel file
#if file becomes to long will mostly likely end up trun in to a class to create a excel object

#USE THIS FILE IF PART OF THE DATA HAS BEEN CONVERTED TO ADC COUNTS 

from openpyxl import Workbook  #Importing the workbook class from the openpyxl library, need to preinstall
from openpyxl import load_workbook as load #need to preinstall
from openpyxl import drawing as ExcelImageInsert
import csv #library already installed with pip
import pandas #need to preinstall
import os #library already installed with pip
import matplotlib.pyplot as plotting #need to preinstall
from  random import randint as Random 

        
#returns difference of ADC counts, fuction will either give a 1D or 2D array
def MaximumCountError(comparison1ideal_data = [], comparison2_data = [], temperature_data = []):
    #comparison data 1 (STRONGER) is mostly 1D, while comparison data (WEAKER) 2 is 2D
    max_temperature_error = 0
    max_temperature_error_array = {'Format': ('max error', 'temperature error occurs at'),'Sensor ' + str(Sensor for Sensor in range(len(comparison2_data))): ()}
    temperature_error_ocurrs_at = 0
    ADC_count_difference = [] #array to collect instance of difference of difference at every temperature

    if(all(isinstance(sublist, list) for sublist in comparison2_data)): #2D data
        for subarray in range(len(comparison2_data)):
            Count_diffence = []
            max_error = 0

            for value in range(len(comparison1ideal_data)): #assume comparison1_data and comparison2_data are the same lenght
                error = (comparison1ideal_data[value] - comparison2_data[subarray][value])
                Count_diffence.append(error)

                if(max_error <= error):
                    max_error = error
                    temperature_error_ocurrs_at = temperature_data[value]

            ADC_count_difference.append(Count_diffence)
            #creating tuple to store maximum error
            max_temperature_error_array['Sensor '+str(subarray)] = (max_error, max_temperature_error)
        
        return ADC_count_difference, max_temperature_error_array
            

    else: #1D data
        for value in range(len(comparison1ideal_data)): #assume comparison1_data and comparison2_data are the same lenght
            error =  (comparison1ideal_data[value]-comparison2_data[value])
            ADC_count_difference.append(error)
            if(max_temperature_error <= error):
                max_temperature_error = error
                temperature_error_ocurrs_at = temperature_data[value]

        return ADC_count_difference, 'The Maximum error between the calibrated data and ideal data) is '+str(max_temperature_error)+' ADC counts.\n At a temperature of '+str(temperature_error_ocurrs_at)+' K'


#Funtions to plot 1D arrays and save graphs
def OneDemensionPlot(x_axis_temperature = [], real_data = [], ideal_data = [], calibrated_data = [] ,graph_label = 'ADC count verses Temperature',x_axis_label = 'Temperature (K)', y_axis_label = 'ADC count/results', real_label = 'Real Data', ideal_label = 'Ideal Data', calibrated_label =  'Calibrated Results', filename = ''):
    if(len(real_data) == len(x_axis_temperature)):
        plotting.plot(x_axis_temperature, real_data, label = real_label)
    if(len(ideal_data) == len(x_axis_temperature)):
        plotting.plot(x_axis_temperature, ideal_data, label = ideal_label)
    if(len(calibrated_data) == len(x_axis_temperature)):
        plotting.plot(x_axis_temperature, calibrated_data, label = calibrated_label)
    plotting.legend()
    plotting.xlabel(x_axis_label)
    plotting.ylabel(y_axis_label)
    plotting.title(graph_label)
    plotting.grid() 
    plotting.savefig(str(filename))
    plotting.close()


#Funtions to plot 2D arrays and save graphs
def TwoDemensionPlot(one_d_x_axis_temperature = [], real_data = [], ideal_data = [], calibrated_data = [] ,graph_label = 'ADC count verses Temperature',x_axis_label = 'Temperature (K)', y_axis_label = 'ADC count/results', real_label = 'Real Data', ideal_label = 'Ideal Data', calibrated_label =  'Calibrated ', filename = '', legend_location = 'upper left', legend_cols = 2, Legend_Outside = False):
    All_data = [real_data, calibrated_data, ideal_data]
    All_data_labels = [real_label, calibrated_label, ideal_label] 
    if(Legend_Outside):
       plotting.figure(figsize=(13,7)) 

    for data in range(len(All_data)):
        if(all(isinstance(sublist, list) for sublist in All_data[data])):
            for subarray in range(len(All_data[data])):
                plotting.plot(one_d_x_axis_temperature, All_data[data][subarray], label = str(All_data_labels[data])+' Sensor '+str(subarray), linestyle='dashed')
        else:
            if(len(All_data[data]) == len(one_d_x_axis_temperature)):
                plotting.plot(one_d_x_axis_temperature, All_data[data], label = str(All_data_labels[data]))

    if(Legend_Outside):
        plotting.legend( loc = legend_location, fontsize = 6, ncol = legend_cols, bbox_to_anchor = (1,1))
        #plotting.tight_layout()
                 
    else:
        plotting.legend( loc = legend_location, fontsize = 6, ncol = legend_cols)
   
    plotting.xlabel(x_axis_label)
    plotting.ylabel(y_axis_label)
    plotting.title(graph_label)
    plotting.grid()
    
    plotting.savefig('Graph_'+str(filename)+'.png', dpi = 250 )
    plotting.close()
    return 'Graph_'+str(filename)+'.png'        


def ConvertExcelToCsv(Excel_name = '', page_number = 1, starting_row = 0, Configure = False):
    reading_excel = pandas.read_excel(Excel_name, sheet_name = None)
    Excel_file = load(Excel_name, read_only=True, keep_links=False)

    #TEMPOARY file formatting and saving
    reading_excel[Excel_file.sheetnames[page_number-1]].to_csv('TESTEXCEL7867.csv')

    #As the first column of the csv file is unneeded, it must be deleted
    #first getting name of column needed for deleting
    Excel_csv = pandas.read_csv('TESTEXCEL7867.csv')
    Excel_csv_column_names = list(Excel_csv.columns)
    #Next deleting column
    del Excel_csv_column_names[0]
    
    if(Configure):
        for rows in range(starting_row):
            Excel_csv = Excel_csv.drop([rows])

    #As the change will only exist as a local change, it must be saves to a new file
    Excel_name = Excel_name.replace('.xlsx','')
    Excel_csv[Excel_csv_column_names].to_csv(Excel_name+'.csv', index=False)
    #deleting temporary csv file, 'TESTEXCEL7867.csv'
    if(os.path.exists('TESTEXCEL7867.csv')):    
        os.remove('TESTEXCEL7867.csv')

    Excel_file.close() #close file to ensure no runtime errors
    return Excel_name+'.csv'


def CSVtoArrayTempandValues(csvfile = csv):
    #Handling the csv files
    cvsArray_storage = []
    temperature_records = []
    Verification_once = True
    
    with open(csvfile) as csv_file:
        next(csv_file) #skips the heading in cvs file
        csv_reader = csv.reader(csv_file) #reading from csv file

        #sorting the csv file into an array to be used in python script
        #due to the nature and arrangment of the excel data will be sort to a 2D array

        for row in csv_reader:
            
            #action needs to happen once
            if(Verification_once):
                cvsArray_storage = [[] for _ in range(len(row)-1)] # assigning a number of arrays to storage array, minus one is to account for the headings in the given csv file

            Verification_once = False
            #+273 K to convert Celcius degrees to Kelvins
            temperature_records.append((float(row[0]))+273.15) #the way the csv file is arranged, places all the temperature values as position 1 (or 0) of each row
            column_counter = 0 #As stated before, each respected value at a given temperature for each sensor is place in the same row, 
            #though more preferable to handle the data of each sensor seperably rather than by temperature

            for value in row:
                if(value != row[0]): #just to avoid collecting any temperature values
                    cvsArray_storage[column_counter].append(float(value))
                    column_counter+=1
            
        #At the end, each array in cvsArray_storage should hold the data for each seperate sensor in a 2D array, also an array for the temperature values
        return temperature_records , cvsArray_storage
    

def DataConversionToADCcount(Array_Vptat = [], Array_Vref = [], ADC_bits = 0, Array_ideal = [], ideal = False):
    ADC_count_conversion = []
   
    #choosing smaller lenght
    if(not ideal):
        Array_lenght_runner = 0 #As the lenghts of the Array could be the same or not, this variable will be choosen to be the smaller lenght of the two to ensure there is noe errrors further on in the code
        
        if(len(Array_Vref) < len(Array_Vptat)): 
            Array_lenght_runner = len(Array_Vref)
        else:
            Array_lenght_runner = len(Array_Vptat) 

        ADC_count_conversion = [[round((Array_Vptat[index][values]/Array_Vref[index][values])*(2**ADC_bits)) for values in range(len(Array_Vptat[index]))] for index in range(Array_lenght_runner)] #calculating ADC count and appending values to new array 
    
    else:
        ADC_count_conversion = [round((Array_ideal[index]/(Array_Vref[index]))*(2**ADC_bits)) for index in range(len(Array_ideal))] #calculating ADC count and appending values to new array

    return  ADC_count_conversion 


#Calibration of real data to ideal data, using the one point calibration method
def OnePointCalibration(Calibration_index = 0, Array_idealADC = [], Array_realADC = [], Temerature_data = []):
    Calibrated_data =  [[round(value*(((Array_idealADC[Calibration_index])/(Temerature_data[Calibration_index]))/((index_race[Calibration_index])/(Temerature_data[Calibration_index])))) for value in index_race] for index_race in Array_realADC]
    
    ''' THE BELOW EQUATION, DESCRIBES THE PROCESS OF THE ABOVE 2D ARRAY
        M_real = (Array_realADC[index_race][Calibration_index])/(Temerature_data[Calibration_index])
        M_ideal = (Array_idealADC[Calibration_index])/(Temerature_data[Calibration_index])
        Data_shift = (M_ideal/M_real) #gtrim_i
        Calibrated_data[index_race].append(round(real_ADC*Data_shift))
    '''
    for index_race in range(len(Array_realADC)): 
        print("All data for sensor "+ str(index_race)+" will be calibrated at "+str(Array_realADC[index_race][Calibration_index])+" for the temperature of "+str(Temerature_data[Calibration_index])+' K')

    return Calibrated_data


#quick functions to wite text and image to excel file
def WriteRandomToExcel(Excel_sheet = Workbook().create_sheet('sheet', 1), row_column = 'B2', paste_data = '', image = False):
    if(not image): Excel_sheet[row_column].value = paste_data
    else:
        graph = ExcelImageInsert.image.Image(paste_data)
        graph.anchor = row_column
        Excel_sheet.add_image(graph)


#deleting orginal graphs
def delete(data):
    if(os.path.exists(data)): 
        os.remove(data)


#function to save excel file
def SaveExcel(workbook = Workbook(), filename = ''):
    #Error can occur if a file is open, this is to prevent programe from failing
    Check_if_file_is_open_validation = True
    Error_Message = 0

    while(Check_if_file_is_open_validation):
        try:
            workbook.save(filename)
            Check_if_file_is_open_validation = False
        except: #if an error occur, this condition will be met
            if(Error_Message == 0):
                print('The file that you wish to save the data to is open, close file to continue.')
                Error_Message+=1


def WriteAndSaveResultsRICExcel(temperature = [], real_array = [], ideal_array = [], calibrated_array = [], filename = ''):
    RIC_workbook = Workbook() #creating Excel file
    Excel_data = ['A_Temperature (Kelvins)', 'B_Real Sensor (ADC Count)', 'C_Ideal Sensor (ADC Count)', 'D_Real(Calibrated) Sensor (ADC Count)'] # A = tempwerature, B = real_array, C = ideal_array, D = calibrated_array

    for sheet_number in range(len(calibrated_array)): # each sensor will have a seperate sheet
        current_sheet = RIC_workbook.create_sheet('Sensor '+str(sheet_number), sheet_number)

        for column in Excel_data:
            #Writing in headings
            current_sheet[column[:1]+str(1)].value = str(column[2:])
            #writing data
            for row in range(2,len(temperature)+1):
                if(column[:1] == 'A'):
                    WriteRandomToExcel(current_sheet, column[:1]+str(row), str(temperature[row-2])) 
                elif(column[:1] == 'B'):
                    WriteRandomToExcel(current_sheet, column[:1]+str(row), str(real_array[sheet_number][row-2]))
                elif(column[:1] == 'C'):
                    WriteRandomToExcel(current_sheet, column[:1]+str(row), str(ideal_array[row-2]))
                else:
                    WriteRandomToExcel(current_sheet, column[:1]+str(row), str(calibrated_array[sheet_number][row-2]))

        #Pasting graph of data to corresponding Excel sheet
        WriteRandomToExcel(current_sheet, 'F3', 'Sensor_Graph_'+str(sheet_number)+'.png', True)
        WriteRandomToExcel(current_sheet, 'F30', 'Error_rate_of_all_Sensor_'+str(sheet_number)+'.png', True)
        _, current_sheet['F29'].value = MaximumCountError(ideal_array, calibrated_array[sheet_number], temperature)

    #in the last sheet
    All_calibrated_to_Ideal = TwoDemensionPlot(temperature, [], ideal_array, calibrated_array, filename = 'Calibrated_and_Ideal')
    WriteRandomToExcel(RIC_workbook.create_sheet('All Sensors', len(calibrated_array)), 'A3', All_calibrated_to_Ideal, True)   
    
    SaveExcel(RIC_workbook, 'Temeperature Sensor Data'+filename+'.xlsx')
    #Deleting orginal file
    for graphs in range(len(calibrated_array)):
        delete('Sensor_Graph_'+str(graphs)+'.png')
        delete('Error_rate_of_all_Sensor_'+str(graphs)+'.png')
    delete(All_calibrated_to_Ideal)

    print("\n-----Excel file of Temperature sensor data has been saved to programme's folder-----\n")


#function to calculate the accuracy of the calibrated data to the ideal data
def Accuracy(Observed_value = [], Actual_value = []):
    Accuracy_rate = [] 
    Error_rate = []

    if(all(isinstance(sublist, list) for sublist in Observed_value)):
        Error_rate = [[(((Observed_value[array][value]-Actual_value[value])/(Actual_value[value]))*100) for value in range(len(Actual_value))] for array in range(len(Observed_value))]
        Accuracy_rate = [[(100 - (abs(Error_rate[array][value]))) for value in range(len(Error_rate[0]))] for array in range(len(Error_rate))]

    else:
        Error_rate = [(((Observed_value[value]-Actual_value[value])/(Actual_value[value]))*100) for value in range(len(Actual_value))]
        Accuracy_rate = [(100 - abs(error)) for error in Error_rate]
        
    return Error_rate, Accuracy_rate


#though as all calibrated data was calibrated to the ideal sensor, the ideal sensor could said to be the line of best fit (maybe)
def LineOfBestFit(x = [], y = []): #clean up
    #y is 1D always

    if(all(isinstance(sublist, list) for sublist in x)):
        x_2D_flip = [[x[row][column] for row in range(len(x))] for column in range(len(x[0]))]
        x = [round(sum(originally_column_now_row)/len(x)) for originally_column_now_row in x_2D_flip]

    x_minus_by_y_minus_sum = sum([((sum(x)/len(x))-x[index])*((sum(y)/len(y))-y[index]) for index in range(len(y))])
    x_minus_square_sum = sum([((sum(x)/len(x))-value)**2 for value in x])

    slope = x_minus_by_y_minus_sum/x_minus_square_sum
    y_intercept = (sum(y)/len(y)) - slope*(sum(x)/len(x))

    return slope, y_intercept


def TemperatureDiffernce(data_major = [], data_minor = [], temperature = []):
    #data_major is the data from sensor 0 : 1D Array
    #data minor is the data all the other sensor bar from sensor 0 : 2D Array(usually)
    count_difference, _ = MaximumCountError(data_major, data_minor, temperature)
    
    index_one = Random(0, len(data_major)-1)
    index_two = Random(0, len(data_major)-1)
    #to ensure the indexs are different
    while(index_one != index_two): index_one = Random(0, len(data_major)-1)

    #for this case the ADC counts will be on the x axis and temperature on the y axis
    slope , _ = LineOfBestFit(data_minor, temperature)
    print('Slope: ',slope)
    
    if (all(isinstance(sublist, list) for sublist in count_difference)): return [[values*slope for values in arrays] for arrays in count_difference]
    return [values*slope for values in count_difference]


#only need to change file names here                   
def main():
    #converting csv files to arrays
    #this while loop is to allow for a fix in error before the prgeam crashes
    Check_if_file_is_open_validation = True
    Error_Message = 0
    while(Check_if_file_is_open_validation):
        try:
            _ , Voltage_ref = CSVtoArrayTempandValues('dts_vref.csv')
            ExcelCSV_realnewname = ConvertExcelToCsv('dts_adc_ADC_count.xlsx', starting_row = 0, Configure = True)
            #for the starting row arguement, you must change the to the row where the labelling begins
            temperature , ADC_sensors_count_results = CSVtoArrayTempandValues(ExcelCSV_realnewname)
            ExcelCSV_idealnewname = ConvertExcelToCsv('VCAL.xlsx')
            _ , Voltage_ideal = CSVtoArrayTempandValues(ExcelCSV_idealnewname)
            Check_if_file_is_open_validation = False
        except: #if an error occur, this condition will be met
            if(Error_Message == 0):
                print('Error has occured, close all relevent files to continue.')
                Error_Message+=1



    #if the code fails to run or has a runtime error, then the above snippet of the main function may be at fault but not the lower bit

    #this section does not need to be altered
    #ADC conversions
    number_of_bits = input('\nEnter the maximum number of output resolution bits for the ADC Count: ')
    while(not number_of_bits.isdigit() or int(number_of_bits) <=  0):
        number_of_bits = input('Invalid input, ensure input is a postive whole number: ')
    number_of_bits = int(number_of_bits)

    index_for_calibration = input('--------------------------------------------------------------------------------------\n\nEnter an index from the real data ADC count that should act as the calibration point: ')
    while(not index_for_calibration.isdigit() or ( 0 >= int(index_for_calibration) or int(index_for_calibration) >= len(temperature)) ):
        index_for_calibration  = input('Invalid input, ensure input is a postive whole number and is between the points of 0 to '+str(len(temperature)-1)+': ')
    index_for_calibration = int(index_for_calibration)
    print('\n--------------------------------------------------------------------------------------\n')    

    #ADC_sensors_count_results = DataConversionToADCcount(Voltage_ptat, Voltage_ref, number_of_bits) #2D Array
    ADC_idealsensor_count_results = DataConversionToADCcount('',Voltage_ref[0] ,number_of_bits, Voltage_ideal[0], True) #1D Array Maybe
    OnePointCalibrationData = OnePointCalibration(index_for_calibration, ADC_idealsensor_count_results, ADC_sensors_count_results, temperature) #2D Array
    Error_array, Accuracy_array = Accuracy(OnePointCalibrationData, ADC_idealsensor_count_results)
    Ideal_Error_Array, Ideal_Accuracy_Array = Accuracy(ADC_idealsensor_count_results, ADC_idealsensor_count_results)    
    
    for index_sensor in range(len(OnePointCalibrationData)):
        OneDemensionPlot(temperature, ADC_sensors_count_results[index_sensor], ADC_idealsensor_count_results, OnePointCalibrationData[index_sensor], filename = 'Sensor_Graph_'+str(index_sensor)+'.png')
        OneDemensionPlot(temperature, [], Ideal_Error_Array, Error_array[index_sensor], 'Error rate for Cailbrated Sensor '+str(index_sensor)+' to Ideal Sensor', y_axis_label = 'Error rate (%)', ideal_label = 'Ideal error rate', calibrated_label = 'Error rate of Cailbrated Sensor '+str(index_sensor), filename = 'Error_rate_of_all_Sensor_'+str(index_sensor))

    TwoDemensionPlot(temperature, [], ADC_idealsensor_count_results, OnePointCalibrationData, filename = 'Calibrated_and_Ideal')
    TwoDemensionPlot(temperature, [], Ideal_Error_Array, Error_array, 'Error rate for all Cailbrated Sensor to Ideal Sensor', y_axis_label = 'Error rate (%)', ideal_label = 'Ideal error rate', calibrated_label = 'Error rate', filename = 'Error_rate_of_all_Sensor', legend_cols = 1, Legend_Outside = True)
    TwoDemensionPlot(temperature, [], Ideal_Accuracy_Array, Accuracy_array, 'Accuracy for all Cailbrated Sensor to Ideal Sensor', y_axis_label = 'Accuracy (%)', ideal_label = 'Ideal Accuracy', calibrated_label = 'Accuracy', filename = 'Accuracy_of_all_Sensor')  
    
    #UNCALIBRATED PLOTS
    #-COUNT ERROR
    Uncalibrated_count_error_ideal, _ = MaximumCountError(ADC_idealsensor_count_results, ADC_sensors_count_results, temperature)
    Ideal_count_error, _ = MaximumCountError(ADC_idealsensor_count_results, ADC_idealsensor_count_results, temperature)
    TwoDemensionPlot(temperature, [], Ideal_count_error, Uncalibrated_count_error_ideal, 'ADC Count error rate for Uncailbrated Sensor to Ideal Sensor', y_axis_label = 'ADC counts', ideal_label = 'Ideal count error', calibrated_label = '', filename = 'Count_error_rate_of_all_Sensor_to_Ideal_UNCALIBRATED',legend_cols = 1, Legend_Outside = True)
    
    Uncalibrated_count_error_sensor_0, _ = MaximumCountError(ADC_sensors_count_results[0], ADC_sensors_count_results, temperature)
    TwoDemensionPlot(temperature, [], [], Uncalibrated_count_error_sensor_0, 'ADC Count error rate for Uncailbrated Sensors to Sensor 0', y_axis_label = 'ADC counts', calibrated_label = '', filename = 'Count_error_rate_of_all_Sensor_to_Sensor_0_UNCALIBRATED', legend_cols = 1, Legend_Outside = True)

    #-TEMPERATURE ERROR
    TwoDemensionPlot(temperature, [], [], TemperatureDiffernce(ADC_sensors_count_results[0], ADC_sensors_count_results[0:], temperature), 'Temperature error rate for Uncailbrated Sensors to Sensor 0', y_axis_label = 'Temperature Error (K)', calibrated_label = '', filename = 'Temperature_error_rate_of_all_Sensor_to_Sensor_0_UNCALIBRATED',legend_cols = 1, Legend_Outside = True)
    

    #CALIBRATED PLOTS
    #-COUNT ERROR
    Calibrated_count_error_sensor_0, _ = MaximumCountError(OnePointCalibrationData[0], OnePointCalibrationData, temperature)
    TwoDemensionPlot(temperature, [], [], Calibrated_count_error_sensor_0, 'ADC Count error rate for Cailbrated Sensors to Sensor 0', y_axis_label = 'ADC counts', calibrated_label = '', filename = 'Count_error_rate_of_all_Sensor_to_Sensor_0_CALIBRATED',legend_cols = 1, Legend_Outside = True)

    #-TEMPERATURE ERROR
    TwoDemensionPlot(temperature, [], [], TemperatureDiffernce(OnePointCalibrationData[0], OnePointCalibrationData[0:], temperature), 'Temperature error rate for Cailbrated Sensors to Sensor 0', y_axis_label = 'Temperature Error (K)', calibrated_label = '', filename = 'Temperature_error_rate_of_all_Sensor_to_Sensor_0_CALIBRATED', legend_cols = 1, Legend_Outside = True)
    
    
    print('\n--------------------------------------------------------------------------------------\n')            
    file_name = input("\nEnter name you wish to assign to file: ")            
    WriteAndSaveResultsRICExcel(temperature, ADC_sensors_count_results, ADC_idealsensor_count_results, OnePointCalibrationData,  filename = str(file_name))
main()    